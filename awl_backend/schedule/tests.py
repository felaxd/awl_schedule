import logging
from datetime import timedelta, date, datetime

import freezegun
from django.core.exceptions import ValidationError
from django.test import TestCase, Client
from django.urls import reverse
from django.utils import timezone
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from common.container import container
from courses.models import Course
from lecturers.models import Lecturer
from rooms.models import Room
from schedule.models import ScheduleBlock, Schedule
from schedule.serializers import ScheduleBlockSerializer
from schedule.services import ExcelScheduleService
from users.models import Group, User


class TestScheduleBlockListViewGet(TestCase):
    def setUp(self):
        self.endpoint = reverse("schedule-block-list")
        self.client = Client()

    def test_all_published_schedule_blocks_will_be_returned(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        response = self.client.get(self.endpoint)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_not_published_schedule_blocks_will_be_filtered(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=False)

        response = self.client.get(self.endpoint)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    @freezegun.freeze_time()
    def test_schedule_blocks_can_be_filtered_by_date_from(self):
        """
        In this case sb1 is in two days from now, so should be returned as we ask for blocks after tomorrow
        Sb2 was finished yesterday so should not be returned
        """
        schedule_block1_start = timezone.now() + timedelta(days=2)
        schedule_block1_end = schedule_block1_start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(
            course_name="Course1", start=schedule_block1_start, end=schedule_block1_end, is_public=True
        )

        schedule_block2_start = timezone.now() - timedelta(days=1)
        schedule_block2_end = schedule_block2_start + timedelta(hours=1)
        schedule_block2 = ScheduleBlock.objects.create(
            course_name="Course2", start=schedule_block2_start, end=schedule_block2_end, is_public=True
        )

        tomorrow = (timezone.now() + timedelta(days=1)).date()
        response = self.client.get(self.endpoint + f"?date_from={tomorrow.isoformat()}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    @freezegun.freeze_time()
    def test_schedule_blocks_can_be_filtered_by_date_to(self):
        """
        In this case sb1 is in two days from now, so should not be returned as we ask for blocks up to tomorrow
        Sb2 was finished yesterday so should be returned
        """
        schedule_block1_start = timezone.now() + timedelta(days=2)
        schedule_block1_end = schedule_block1_start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(
            course_name="Course1", start=schedule_block1_start, end=schedule_block1_end, is_public=True
        )

        schedule_block2_start = timezone.now() - timedelta(days=1)
        schedule_block2_end = schedule_block2_start + timedelta(hours=1)
        schedule_block2 = ScheduleBlock.objects.create(
            course_name="Course2", start=schedule_block2_start, end=schedule_block2_end, is_public=True
        )

        tomorrow = (timezone.now() + timedelta(days=1)).date()
        response = self.client.get(self.endpoint + f"?date_to={tomorrow.isoformat()}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    @freezegun.freeze_time()
    def test_schedule_blocks_can_be_filtered_by_timerange(self):
        """In this case sb1 and sb2 are in date range. Sb3 is a week after timerange so should be filtered out"""
        schedule_block1_start = timezone.now() + timedelta(days=2)
        schedule_block1_end = schedule_block1_start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(
            course_name="Course1", start=schedule_block1_start, end=schedule_block1_end, is_public=True
        )

        schedule_block2_start = timezone.now() - timedelta(days=1)
        schedule_block2_end = schedule_block2_start + timedelta(hours=1)
        schedule_block2 = ScheduleBlock.objects.create(
            course_name="Course2", start=schedule_block2_start, end=schedule_block2_end, is_public=True
        )

        schedule_block3_start = timezone.now() - timedelta(weeks=2)
        schedule_block3_end = schedule_block2_start + timedelta(hours=1)
        schedule_block3 = ScheduleBlock.objects.create(
            course_name="Course2", start=schedule_block3_start, end=schedule_block3_end, is_public=True
        )

        week_start = (timezone.now() - timedelta(weeks=1)).date()
        week_end = (timezone.now() + timedelta(weeks=1)).date()
        response = self.client.get(
            self.endpoint + f"?date_from={week_start.isoformat()}&date_to={week_end.isoformat()}"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertIn(ScheduleBlockSerializer(schedule_block2).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block3).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_lecturer(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        lecturer = Lecturer.objects.create(first_name="Lecturer", last_name="1", is_public=True)
        schedule_block1.lecturers.add(lecturer, through_defaults={"room": None})

        response = self.client.get(self.endpoint + f"?lecturers={lecturer.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_multiple_lecturers(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        lecturer1 = Lecturer.objects.create(first_name="Lecturer", last_name="1", is_public=True)
        schedule_block1.lecturers.add(lecturer1, through_defaults={"room": None})

        lecturer2 = Lecturer.objects.create(first_name="Lecturer", last_name="2", is_public=True)
        schedule_block2.lecturers.add(lecturer1, through_defaults={"room": None})

        response = self.client.get(self.endpoint + f"?lecturers={lecturer1.id}&lecturers={lecturer2.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_room(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        room = Room.objects.create(name="Room1", is_public=True)
        schedule_block1.rooms.add(room)

        response = self.client.get(self.endpoint + f"?rooms={room.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_multiple_rooms(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        room1 = Room.objects.create(name="Room1", is_public=True)
        schedule_block1.rooms.add(room1)

        room2 = Room.objects.create(name="Room2", is_public=True)
        schedule_block2.rooms.add(room2)

        response = self.client.get(self.endpoint + f"?rooms={room1.id}&rooms={room2.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_group(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        group = Group.objects.create(name="Group1", is_public=True)
        schedule_block1.groups.add(group)

        response = self.client.get(self.endpoint + f"?groups={group.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_multiple_groups(self):
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        group1 = Group.objects.create(name="Group1", is_public=True)
        schedule_block1.groups.add(group1)

        group2 = Group.objects.create(name="Group2", is_public=True)
        schedule_block2.groups.add(group2)

        response = self.client.get(self.endpoint + f"?groups={group1.id}&groups={group2.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 2)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_mix_of_parameters_case1(self):
        """
        Filter by room and group
        Sb1 has correct group and room
        Sb2 has only correct room
        """
        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        schedule_block2 = ScheduleBlock.objects.create(course_name="Course2", start=start, end=end, is_public=True)

        group = Group.objects.create(name="Group1", is_public=True)
        schedule_block1.groups.add(group)

        room = Room.objects.create(name="Room1", is_public=True)
        schedule_block1.rooms.add(room)
        schedule_block2.rooms.add(room)

        response = self.client.get(self.endpoint + f"?groups={group.id}&rooms={room.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)

    def test_schedule_blocks_can_be_filtered_by_mix_of_parameters_case2(self):
        """
        Filter by group and timerange
        Sb1 has correct group and is in timerange
        Sb2 is in correct timerange, but has no requested group
        Sb3 has correct group but is not in timerange
        """
        schedule_block1_start = timezone.now() + timedelta(days=2)
        schedule_block1_end = schedule_block1_start + timedelta(hours=1)
        schedule_block1 = ScheduleBlock.objects.create(
            course_name="Course1", start=schedule_block1_start, end=schedule_block1_end, is_public=True
        )

        schedule_block2_start = timezone.now() - timedelta(days=1)
        schedule_block2_end = schedule_block2_start + timedelta(hours=1)
        schedule_block2 = ScheduleBlock.objects.create(
            course_name="Course2", start=schedule_block2_start, end=schedule_block2_end, is_public=True
        )

        schedule_block3_start = timezone.now() - timedelta(weeks=2)
        schedule_block3_end = schedule_block2_start + timedelta(hours=1)
        schedule_block3 = ScheduleBlock.objects.create(
            course_name="Course2", start=schedule_block3_start, end=schedule_block3_end, is_public=True
        )

        group = Group.objects.create(name="Group1", is_public=True)
        schedule_block1.groups.add(group)
        schedule_block3.groups.add(group)

        week_start = (timezone.now() - timedelta(weeks=1)).date()
        week_end = (timezone.now() + timedelta(weeks=1)).date()
        response = self.client.get(
            self.endpoint + f"?groups={group.id}&date_from={week_start.isoformat()}&date_to={week_end.isoformat()}"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)
        self.assertIn(ScheduleBlockSerializer(schedule_block1).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block2).data, response.data)
        self.assertNotIn(ScheduleBlockSerializer(schedule_block3).data, response.data)


class TestExcelScheduleServiceFile(TestCase):
    def setUp(self):
        self.service = ExcelScheduleService()
        self.path = "schedule/test_cassettes/main_template.xlsx"
        self.workbook = self.service.load_workbook(self.path)
        self.worksheet = self.service.load_worksheet(self.workbook, "TEMPLATE")
        self.service.current_workbook = self.workbook

    def test_can_load_workbook(self):
        self.assertIsInstance(self.workbook, Workbook)

    def test_can_list_workbook_worksheets(self):
        worksheets = self.service.get_excel_file_worksheet_list(self.path)
        self.assertEqual(worksheets, ["TEMPLATE"])

    def test_can_load_worksheet(self):
        self.assertIsInstance(self.worksheet, Worksheet)

    def test_can_get_worksheet_cell(self):
        cell = self.service.get_cell(self.worksheet, row=1, column=2)
        self.assertIsInstance(cell, Cell)
        self.assertEqual(cell.value, "STUDIA CYWILNE - WYDZIAŁ ZARZĄDZANIA")

    def test_init_excel_worksheet(self):
        logging.disable(logging.WARNING)
        result = self.service.init_excel_worksheet(
            path=self.path, worksheet="TEMPLATE", year=2023, month=10
        )
        self.assertIsInstance(result["workbook"], Workbook)
        self.assertIsInstance(result["worksheet"], Worksheet)
        self.assertEqual(result["year"], 2023)
        self.assertEqual(result["month"], 10)

        schedule_days = list(filter(lambda day: "error" not in day, result["schedule_days"]))
        self.assertEqual(len(schedule_days), 1)

    def test_init_excel_worksheet_raises_error_when_wrong_date_is_given(self):
        with self.assertRaisesMessage(ValidationError, "Brak dni w arkuszu dla podenego miesiąca"):
            self.service.init_excel_worksheet(
                path=self.path, worksheet="TEMPLATE", year=2024, month=10
            )

    def test_get_single_day_schedule_info(self):
        logging.disable()
        result = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=1
        )
        self.assertEqual(result["date"], date(2023, 10, 1))
        self.assertEqual(result["starting_cell"].coordinate, "A1")
        self.assertEqual(result["ending_cell"].coordinate, "Q11")
        self.assertEqual(len(result["groups"]), 1)
        self.assertIn({"name": "GROUP1", "start_row": 6, "end_row": 11}, result["groups"])
        excluded_cells = [cell.coordinate for cell in result["excluded_cells"]]
        self.assertEqual(len(excluded_cells), 85)  # top 5 rows x 17 columns
        # test some cells
        for cell in ["A1", "A2", "A3", "A4", "A5"]:
            self.assertIn(cell, excluded_cells)


class TestExcelScheduleServiceModuleInfo(TestCase):
    def setUp(self):
        self.service = ExcelScheduleService()
        self.path = "schedule/test_cassettes/schedule_with_test_cases.xlsx"
        self.workbook = self.service.load_workbook(self.path)
        self.worksheet = self.service.load_worksheet(self.workbook, "TEST_CASES")
        self.service.current_workbook = self.workbook

    def test_get_module_info(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=2
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 19)
        module_info = self.service.get_module_info(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(module_info["name"], "MATEMATYKA")
        self.assertEqual(module_info["type"], "wykład")
        self.assertEqual(module_info["start"], datetime(2023, 10, 2, 8, 0, 0))
        self.assertEqual(module_info["end"], datetime(2023, 10, 2, 9, 30, 0))
        self.assertEqual(len(module_info["lecturers"]), 2)
        self.assertEqual(len(module_info["rooms"]), 2)
        self.assertEqual(len(module_info["groups"]), 1)
        self.assertEqual(module_info["starting_cell"].coordinate, "S6")
        self.assertEqual(module_info["ending_cell"].coordinate, "T11")

    def test_get_module_name(self):
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 2)
        module_name = self.service.get_module_name(module_starting_cell)
        self.assertEqual(module_name, "MATEMATYKA")

    def test_get_module_type_from_name(self):
        module_name = self.service.get_module_type_from_name("MATEMATYKA -Ć")
        self.assertEqual(module_name, "ćwiczenia")

        module_name = self.service.get_module_type_from_name("MATEMATYKA -L")
        self.assertEqual(module_name, "laboratorium")

        module_name = self.service.get_module_type_from_name("MATEMATYKA -F")
        self.assertEqual(module_name, "fakultet")

    def test_get_module_lecturers_single(self):
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 2)
        module_ending_cell = self.service.get_cell(self.worksheet, 11, 3)
        lecturers = self.service.get_module_lecturers(
            self.worksheet, module_starting_cell, module_ending_cell
        )
        self.assertEqual(lecturers[0], {"name": "KOWALSKI", "room": "2.003"})

    def test_get_module_lecturers_multiple(self):
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 19)
        module_ending_cell = self.service.get_cell(self.worksheet, 11, 20)
        lecturers = self.service.get_module_lecturers(
            self.worksheet, module_starting_cell, module_ending_cell
        )
        self.assertEqual(len(lecturers), 2)
        self.assertIn({"name": "KOWALSKI", "room": "2.003"}, lecturers)
        self.assertIn({"name": "NOWAK", "room": "2.004"}, lecturers)

    def test_get_module_rooms_single(self):
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 2)
        module_ending_cell = self.service.get_cell(self.worksheet, 11, 3)
        rooms = self.service.get_module_rooms(
            self.worksheet, module_starting_cell, module_ending_cell
        )
        self.assertEqual(["2.003"], rooms)

    def test_get_module_rooms_multiple(self):
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 19)
        module_ending_cell = self.service.get_cell(self.worksheet, 11, 20)
        rooms = self.service.get_module_rooms(
            self.worksheet, module_starting_cell, module_ending_cell
        )
        self.assertEqual(len(rooms), 2)
        self.assertEqual(["2.003", "2.004"], rooms)

    def test_get_module_groups_single(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=1
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 2)
        module_ending_cell = self.service.get_cell(self.worksheet, 11, 3)
        groups = self.service.get_module_groups(
            single_day_schedule_info["groups"], module_starting_cell, module_ending_cell
        )
        self.assertEqual(["GRUPA1"], groups)

    def test_get_module_groups_multiple(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=4
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 53)
        module_ending_cell = self.service.get_cell(self.worksheet, 17, 54)
        groups = self.service.get_module_groups(
            single_day_schedule_info["groups"], module_starting_cell, module_ending_cell
        )
        self.assertEqual(len(groups), 2)
        self.assertEqual(["GRUPA1", "GRUPA2"], groups)

    def test_get_module_colour(self):
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 2)
        colour = self.service.get_module_colour(module_starting_cell)
        self.assertEqual(colour, "#FFC4BD97")

    def test_get_module_range(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=1
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 2)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "B6")
        self.assertEqual(e_cell.coordinate, "C11")

    def test_get_module_range_extended_horizontally(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=3
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 36)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "AJ6")
        self.assertEqual(e_cell.coordinate, "AM11")

    def test_get_module_range_extended_vertically(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=4
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 53)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "BA6")
        self.assertEqual(e_cell.coordinate, "BB17")

    def test_get_module_range_no_colours(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=11
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 172)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "FP6")
        self.assertEqual(e_cell.coordinate, "FQ11")

    def test_get_module_range_broken_colours(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=10
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 155)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "EY6")
        self.assertEqual(e_cell.coordinate, "EZ11")

    def test_get_module_range_broken_borders(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=6
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 87)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "CI6")
        self.assertEqual(e_cell.coordinate, "CJ11")

    def test_get_module_range_broken_borders_between_schedule_blocks(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(
            worksheet=self.worksheet, year=2023, month=10, day=12
        )
        module_starting_cell = self.service.get_cell(self.worksheet, 6, 189)
        s_cell, e_cell = self.service.get_module_range(
            single_day_schedule_info, module_starting_cell
        )
        self.assertEqual(s_cell.coordinate, "GG6")
        self.assertEqual(e_cell.coordinate, "GH11")


class TestExcelScheduleServiceGetSchedule(TestCase):
    def setUp(self):
        self.service = ExcelScheduleService()
        self.path = "schedule/test_cassettes/schedule_with_test_cases.xlsx"
        self.workbook = self.service.load_workbook(self.path)
        self.worksheet = self.service.load_worksheet(self.workbook, "TEST_CASES")
        self.service.current_workbook = self.workbook

    def test_get_schedule_for_single_day(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(self.worksheet, 2023, 10, 7)
        schedule = self.service.get_schedule_for_single_day(single_day_schedule_info)

        sb1_module_info = {
            'name': 'MATEMATYKA',
            'type': 'wykład',
            'start': datetime(2023, 10, 7, 8, 0),
            'end': datetime(2023, 10, 7, 9, 30),
            'lecturers': [{'name': 'KOWALSKI', 'room': '2.003'}],
            'rooms': ['2.003'],
            'groups': ['GRUPA1'],
            'colour': '#FFC4BD97',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 104),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 105)
        }

        sb2_module_info = {
            'name': 'MIKROEKONOMIA -Ć',
            'type': 'ćwiczenia',
            'start': datetime(2023, 10, 7, 9, 40),
            'end': datetime(2023, 10, 7, 11, 10),
            'lecturers': [{'name': 'NOWAK', 'room': '2.002'}],
            'rooms': ['2.002'],
            'groups': ['GRUPA1'],
            'colour': '#00F8CBAD',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 106),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 107)
        }
        sb3_module_info = {
            'name': 'KOMUNIKACJA W ZARZĄDZANIU',
            'type': 'wykład',
            'start': datetime(2023, 10, 7, 13, 10),
            'end': datetime(2023, 10, 7, 14, 40),
            'lecturers': [{'name': 'WIŚNIEWSKA', 'room': '4.039'}],
            'rooms': ['4.039'],
            'groups': ['GRUPA1'],
            'colour': '#00C5E0B4',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 110),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 111)
        }

        self.assertEqual(len(schedule), 3)
        self.assertIn(sb1_module_info, schedule)
        self.assertIn(sb2_module_info, schedule)
        self.assertIn(sb3_module_info, schedule)

    def test_get_schedule_for_single_day_multiple_groups(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(self.worksheet, 2023, 10, 5)
        schedule = self.service.get_schedule_for_single_day(single_day_schedule_info)

        sb1_module_info = {
            'name': 'MATEMATYKA',
            'type': 'wykład',
            'start': datetime(2023, 10, 5, 8, 0),
            'end': datetime(2023, 10, 5, 9, 30),
            'lecturers': [{'name': 'KOWALSKI', 'room': '2.003'}],
            'rooms': ['2.003'],
            'groups': ['GRUPA1'],
            'colour': '#FFC4BD97',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 70),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 71)
        }

        sb2_module_info = {
            'name': 'MIKROEKONOMIA',
            'type': 'wykład',
            'start': datetime(2023, 10, 5, 11, 30),
            'end': datetime(2023, 10, 5, 13, 00),
            'lecturers': [{'name': 'NOWAK', 'room': '2.002'}],
            'rooms': ['2.002'],
            'groups': ['GRUPA2'],
            'colour': '#00F8CBAD',
            'starting_cell': self.service.get_cell(self.worksheet, 12, 74),
            'ending_cell': self.service.get_cell(self.worksheet, 17, 75)
        }

        self.assertEqual(len(schedule), 2)
        self.assertIn(sb1_module_info, schedule)
        self.assertIn(sb2_module_info, schedule)

    def test_get_schedule_for_single_day_broken_borders(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(self.worksheet, 2023, 10, 8)
        schedule = self.service.get_schedule_for_single_day(single_day_schedule_info)

        sb1_module_info = {
            'name': 'PODSTAWY PEDAGOGIKI',
            'type': 'wykład',
            'start': datetime(2023, 10, 8, 8, 0),
            'end': datetime(2023, 10, 8, 9, 30),
            'lecturers': [{'name': 'KOWALSKA', 'room': '2.003'}],
            'rooms': ['2.003'],
            'groups': ['GRUPA1'],
            'colour': '#FFC4BD97',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 121),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 122)
        }

        sb2_module_info = {
            'name': 'WF',
            'type': 'wykład',
            'start': datetime(2023, 10, 8, 9, 40),
            'end': datetime(2023, 10, 8, 11, 10),
            'lecturers': [{'name': 'KOWALCZYK', 'room': 'SALA'}],
            'rooms': ['SALA'],
            'groups': ['GRUPA1'],
            'colour': '#00C5E0B4',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 123),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 124)
        }
        sb3_module_info = {
            'name': 'APEL',
            'type': 'wykład',
            'start': datetime(2023, 10, 8, 11, 30),
            'end': datetime(2023, 10, 8, 14, 40),
            'lecturers': [],
            'rooms': ['2.003'],
            'groups': ['GRUPA1'],
            'colour': '#FFC4BD97',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 125),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 128)
        }

        self.assertEqual(len(schedule), 3)
        self.assertIn(sb1_module_info, schedule)
        self.assertIn(sb2_module_info, schedule)
        self.assertIn(sb3_module_info, schedule)

    def test_get_schedule_for_single_day_broken_colours(self):
        single_day_schedule_info = self.service.get_single_day_schedule_info(self.worksheet, 2023, 10, 9)
        schedule = self.service.get_schedule_for_single_day(single_day_schedule_info)

        sb1_module_info = {
            'name': 'MATEMATYKA',
            'type': 'wykład',
            'start': datetime(2023, 10, 9, 8, 0),
            'end': datetime(2023, 10, 9, 9, 30),
            'lecturers': [{'name': 'KOWALSKI', 'room': '2.003'}],
            'rooms': ['2.003'],
            'groups': ['GRUPA1'],
            'colour': '#FFC4BD97',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 138),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 139)
        }

        sb2_module_info = {
            'name': 'MIKROEKONOMIA -Ć',
            'type': 'ćwiczenia',
            'start': datetime(2023, 10, 9, 9, 40),
            'end': datetime(2023, 10, 9, 11, 10),
            'lecturers': [{'name': 'NOWAK', 'room': '2.002'}],
            'rooms': ['2.002'],
            'groups': ['GRUPA1'],
            'colour': '#00F8CBAD',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 140),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 141)
        }
        sb3_module_info = {
            'name': 'KOMUNIKACJA W ZARZĄDZANIU',
            'type': 'wykład',
            'start': datetime(2023, 10, 9, 13, 10),
            'end': datetime(2023, 10, 9, 14, 40),
            'lecturers': [{'name': 'WIŚNIEWSKA', 'room': '4.039'}],
            'rooms': ['4.039'],
            'groups': ['GRUPA1'],
            'colour': '#00C5E0B4',
            'starting_cell': self.service.get_cell(self.worksheet, 6, 144),
            'ending_cell': self.service.get_cell(self.worksheet, 11, 145)
        }

        self.assertEqual(len(schedule), 3)
        self.assertIn(sb1_module_info, schedule)
        self.assertIn(sb2_module_info, schedule)
        self.assertIn(sb3_module_info, schedule)


class TestScheduleServiceUpdate(TestCase):
    def setUp(self):
        logging.disable()

    def test_update_schedule_from_excel(self):
        # self.skipTest(reason="For speeding up test process test was skipped")
        schedule = Schedule.objects.create(
            creator=User.objects.create_user("test_user", "test@test.django.com"),
            name="TEST",
            file="schedule/test_cassettes/full_schedule.xlsx",
            year=2023,
            month=10,
            worksheet_name="PAŹDZIERNIK"
        )
        result = container().schedule_service.update_schedule_from_excel(schedule)

        self.assertEqual(len(result["replaced_blocks"]), 0)
        self.assertEqual(len(result["added_blocks"]), 1076)
        self.assertEqual(len(result["added_groups"]), 23)
        self.assertEqual(len(result["added_lecturers"]), 125)
        self.assertEqual(len(result["added_rooms"]), 59)
        self.assertEqual(len(result["added_courses"]), 352)
        self.assertEqual(len(result["errors"]), 0)

        self.assertEqual(ScheduleBlock.objects.count(), 1076)
        self.assertEqual(Group.objects.count(), 23)
        self.assertEqual(Lecturer.objects.count(), 125)
        self.assertEqual(Room.objects.count(), 59)
        self.assertEqual(Course.objects.count(), 352)

    def test_publicate_schedule(self):
        schedule = Schedule.objects.create(
            creator=User.objects.create_user("test_user", "test@test.django.com"),
            name="TEST",
            file="schedule/test_cassettes/full_schedule.xlsx",
            year=2023,
            month=10,
            worksheet_name="PAŹDZIERNIK",
            status="FINISHED",
        )

        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=False)
        room = Room.objects.create(name="Room1", is_public=False)
        course = Course.objects.create(name="Course1", is_public=False)
        group = Group.objects.create(name="Group1", is_public=False)
        lecturer = Lecturer.objects.create(first_name="Lecturer", last_name="1", is_public=False)

        schedule.schedule_blocks.add(schedule_block)
        schedule.rooms.add(room)
        schedule.courses.add(course)
        schedule.groups.add(group)
        schedule.lecturers.add(lecturer)

        self.assertFalse(ScheduleBlock.objects.filter(is_public=True).exists())
        self.assertFalse(Group.objects.filter(is_public=True).exists())
        self.assertFalse(Lecturer.objects.filter(is_public=True).exists())
        self.assertFalse(Room.objects.filter(is_public=True).exists())
        self.assertFalse(Course.objects.filter(is_public=True).exists())

        container().schedule_service.publicate_schedule(schedule)

        self.assertTrue(ScheduleBlock.objects.filter(is_public=True).exists())
        self.assertTrue(Group.objects.filter(is_public=True).exists())
        self.assertTrue(Lecturer.objects.filter(is_public=True).exists())
        self.assertTrue(Room.objects.filter(is_public=True).exists())
        self.assertTrue(Course.objects.filter(is_public=True).exists())

    def test_revert_schedule_publication(self):
        schedule = Schedule.objects.create(
            creator=User.objects.create_user("test_user", "test@test.django.com"),
            name="TEST",
            file="schedule/test_cassettes/full_schedule.xlsx",
            year=2023,
            month=10,
            worksheet_name="PAŹDZIERNIK",
            status="PUBLICATED",
        )

        start = timezone.now()
        end = start + timedelta(hours=1)
        schedule_block = ScheduleBlock.objects.create(course_name="Course1", start=start, end=end, is_public=True)
        room = Room.objects.create(name="Room1", is_public=True)
        course = Course.objects.create(name="Course1", is_public=True)
        group = Group.objects.create(name="Group1", is_public=True)
        lecturer = Lecturer.objects.create(first_name="Lecturer", last_name="1", is_public=True)

        schedule.schedule_blocks.add(schedule_block)
        schedule.rooms.add(room)
        schedule.courses.add(course)
        schedule.groups.add(group)
        schedule.lecturers.add(lecturer)

        self.assertTrue(ScheduleBlock.objects.filter(is_public=True).exists())
        self.assertTrue(Group.objects.filter(is_public=True).exists())
        self.assertTrue(Lecturer.objects.filter(is_public=True).exists())
        self.assertTrue(Room.objects.filter(is_public=True).exists())
        self.assertTrue(Course.objects.filter(is_public=True).exists())

        container().schedule_service.revert_schedule_publication(schedule)

        self.assertFalse(ScheduleBlock.objects.filter(is_public=True).exists())
        self.assertFalse(Group.objects.filter(is_public=True).exists())
        self.assertFalse(Lecturer.objects.filter(is_public=True).exists())
        self.assertFalse(Room.objects.filter(is_public=True).exists())
        self.assertFalse(Course.objects.filter(is_public=True).exists())
