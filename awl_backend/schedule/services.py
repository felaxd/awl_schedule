import calendar
import logging
from dataclasses import dataclass
from datetime import datetime, date
from typing import Dict, Any, Optional, List, Union, Tuple
from uuid import UUID

from django.db.models import F
from openpyxl import Workbook

from django.contrib.admin.models import ADDITION, CHANGE, DELETION
from django.core.exceptions import ValidationError
from django.db import transaction, IntegrityError
from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Side
from openpyxl.worksheet.worksheet import Worksheet

from common.utils import django_log_action
from common.validators import validate_user_permission
from courses.models import Course
from lecturers.models import Lecturer
from rooms.models import Room
from schedule.consts import AMOUNT_OF_TIME_BLOCK_PER_DAY, HOUR_BLOCK_START_TIMESPANS, DEFAULT_MODULE_TYPE, \
    MODULE_TYPES_TUPLE
from schedule.excel_colours import theme_and_tint_to_rgb
from schedule.models import ScheduleBlock, Schedule, LecturerScheduleBlockThrough
from schedule.selectors import ScheduleBlockSelector
from users.models import User, Group

log = logging.getLogger("schedule")


@dataclass
class ScheduleBlockService:

    @staticmethod
    @transaction.atomic()
    def set_schedule_block_lecturers(
        user: User, schedule_block: ScheduleBlock, lecturers: List[Dict[str, Union[Lecturer, Room]]]
    ) -> None:
        schedule_block.lecturers.clear()
        for lecturer in lecturers:
            _lecturer = lecturer.get("lecturer")
            schedule_block.lecturers.add(_lecturer, through_defaults={"room": lecturer.get("room", None)})

        change_message = f"Changed lecturers ({', '.join([l.__str__() for l in schedule_block.lecturers.all()])})"
        django_log_action(user=user, obj=schedule_block, action_flag=CHANGE, change_message=change_message)

    @staticmethod
    @transaction.atomic()
    def set_schedule_block_groups(user: User, schedule_block: ScheduleBlock, groups: List[Group]) -> None:
        schedule_block.groups.clear()
        for group in groups:
            schedule_block.groups.add(group)
        change_message = f"Changed groups ({', '.join(schedule_block.groups.values_list('name', flat=True))})"
        django_log_action(user=user, obj=schedule_block, action_flag=CHANGE, change_message=change_message)

    @staticmethod
    @transaction.atomic()
    def set_schedule_block_rooms(user: User, schedule_block: ScheduleBlock, rooms: List[Room]) -> None:
        schedule_block.rooms.clear()
        for room in rooms:
            schedule_block.rooms.add(room)
        change_message = f"Changed rooms ({', '.join(schedule_block.rooms.values_list('name', flat=True))})"
        django_log_action(user=user, obj=schedule_block, action_flag=CHANGE, change_message=change_message)

    @transaction.atomic()
    def create_schedule_block(
        self,
        user: User,
        course_name: str,
        course: Optional[Course],
        start: datetime,
        end: datetime,
        type: str,
        lecturers: Optional[List[Dict[str, Union[Lecturer, Room]]]] = None,
        groups: Optional[List[Group]] = None,
        rooms: Optional[List[Room]] = None,
        colour: Optional[str] = None,
    ) -> ScheduleBlock:
        """Creates schedule block"""
        schedule_block = ScheduleBlock(
            course_name=course_name,
            course=course,
            start=start,
            end=end,
            type=type,
            colour=colour or "",
        )
        validate_user_permission(user, schedule_block.ADD_PERMISSION_CODENAME, raise_error=True)
        schedule_block.full_clean()
        schedule_block.save()

        django_log_action(user=user, obj=schedule_block, action_flag=ADDITION)

        if lecturers:
            self.set_schedule_block_lecturers(user, schedule_block, lecturers)

        if groups:
            self.set_schedule_block_groups(user, schedule_block, groups)
        else:
            raise ValidationError({"groups": "Lista nie może być pusta."})

        if rooms:
            self.set_schedule_block_rooms(user, schedule_block, rooms)

        return schedule_block

    @transaction.atomic()
    def update_schedule_block(self, user: User, schedule_block_id: UUID, **kwargs: Dict[str, Any]) -> ScheduleBlock:
        """Updates schedule block details"""
        schedule_block = ScheduleBlockSelector.get_by_id(schedule_block_id)
        validate_user_permission(user, schedule_block.CHANGE_PERMISSION_CODENAME, raise_error=True)

        lecturers: Optional[List[Dict[str, Union[Lecturer, Room]]]] = kwargs.pop("lecturers", None)
        groups: Optional[List[Group]] = kwargs.pop("groups", None)
        rooms: Optional[List[Room]] = kwargs.pop("rooms", None)

        # serialize input so kwargs has only Lecturer object fields
        model_fields = [field.name for field in ScheduleBlock._meta.fields]
        if kwargs := {k: kwargs[k] for k in model_fields if k in kwargs.keys()}:
            ScheduleBlock.objects.filter(id=schedule_block.id).update(**kwargs)
            schedule_block.refresh_from_db()
            schedule_block.full_clean()

            changed = schedule_block.changed_fields
            if changed:
                change_message = [{"changed": {"fields": changed}}]
                django_log_action(user=user, obj=schedule_block, action_flag=CHANGE, change_message=change_message)

        if lecturers is not None:
            self.set_schedule_block_lecturers(user, schedule_block, lecturers)

        if groups is not None:
            self.set_schedule_block_groups(user, schedule_block, groups)

        if rooms is not None:
            self.set_schedule_block_rooms(user, schedule_block, rooms)

        return schedule_block

    @staticmethod
    @transaction.atomic()
    def delete_schedule_block(user: User, schedule_block_id: UUID) -> None:
        schedule_block = ScheduleBlockSelector.get_by_id(schedule_block_id)
        validate_user_permission(user, schedule_block.DELETE_PERMISSION_CODENAME, raise_error=True)
        schedule_block.delete()

        django_log_action(user=user, obj=schedule_block, action_flag=DELETION)
        return None


@dataclass
class ExcelScheduleService:
    current_workbook = None

    @staticmethod
    def load_workbook(path: str) -> Workbook:
        try:
            return load_workbook(filename=path, data_only=True)
        except Exception as ex:
            raise ValidationError(f"Błąd podczas odczytu pliku excel: {ex}")

    @staticmethod
    def load_worksheet(workbook: Workbook, worksheet: str) -> Worksheet:
        try:
            return workbook[worksheet]
        except KeyError:
            raise ValidationError("Arkusz nie istnieje w pliku")

    @staticmethod
    def get_cell(worksheet: Worksheet, row: int, column: int) -> Cell:
        return worksheet.cell(row=row, column=column)

    def get_excel_file_worksheet_list(self, path: str) -> List[str]:
        wb = self.load_workbook(path)
        return wb.sheetnames

    def init_excel_worksheet(self, path: str, worksheet: str, year: int, month: int) -> Dict[str, Any]:
        wb = self.load_workbook(path)
        self.current_workbook = wb
        ws = self.load_worksheet(wb, worksheet)
        schedule_days = []
        added_days = 0
        last_day_of_month = calendar.monthrange(year, month)[1]
        for day in range(1, last_day_of_month):
            try:
                schedule_days.append(self.get_single_day_schedule_info(ws, year, month, day))
                added_days += 1
            except ValidationError as ex:
                error = {"date": date(day=day, month=month, year=year), "error": f"{ex}"}
                log.warning(error)
                schedule_days.append(error)

        if not added_days:
            raise ValidationError("Brak dni w arkuszu dla podenego miesiąca")

        return {
            "workbook": wb,
            "worksheet": ws,
            "year": year,
            "month": month,
            "schedule_days": schedule_days,
        }

    def get_single_day_schedule_info(self, worksheet: Worksheet, year: int, month: int, day: int) -> Dict[str, Any]:
        _date = datetime(day=day, month=month, year=year)
        log.debug(f"Looking for day {_date.date()}")

        date_cell = self.get_cell_with_value(worksheet, _date, exact=True)

        if not date_cell:
            raise ValidationError("Brak daty w arkuszu")

        starting_cell = self.get_cell(worksheet=worksheet, row=date_cell.row - 1, column=date_cell.column - 4)
        ending_cell = self.find_lower_boundary_of_column(worksheet, date_cell.column + 12)
        log.debug({"starting_cell": starting_cell, "ending_cell": ending_cell})

        excluded_cells = []

        groups = []
        groups_column = starting_cell.column
        group_rows: List[Tuple[Cell]] = worksheet.iter_rows(
            min_col=groups_column,
            max_col=groups_column,
            min_row=starting_cell.row,
            max_row=ending_cell.row
        )
        group_name_parts = []
        last_group_start = starting_cell.row
        for row in group_rows:
            cell = row[0]
            cell_under = self.get_cell(worksheet, row=cell.row + 1, column=groups_column)
            if value := cell.value:
                group_name_parts.append(f"{value}")
            if cell.border.bottom.style or cell_under.border.top.style:
                group_name_parts = list(filter(None, group_name_parts))
                group_name = " ".join(group_name_parts)
                parsed_group_name = group_name.replace(" ", "")
                if len(parsed_group_name) <= 1:
                    for row_in_last_range in range(last_group_start, cell.row + 1):
                        cells_to_exclude = worksheet.iter_cols(
                            min_col=starting_cell.column,
                            max_col=ending_cell.column,
                            min_row=row_in_last_range,
                            max_row=row_in_last_range
                        )
                        for cell_to_exclude in cells_to_exclude:
                            if cell_to_exclude[0] not in excluded_cells:
                                excluded_cells.append(cell_to_exclude[0])
                else:
                    _group = {
                        "name": group_name,
                        "start_row": last_group_start,
                        "end_row": cell.row
                    }
                    log.debug(f"Group found: {_group}")
                    groups.append(_group)
                group_name_parts = []
                last_group_start = cell.row + 1

        return {
            "worksheet": worksheet,
            "date": _date.date(),
            "starting_cell": starting_cell,
            "ending_cell": ending_cell,
            "groups": groups,
            "excluded_cells": excluded_cells,
        }

    def get_schedule_for_single_day(self, single_day_schedule_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        worksheet: Worksheet = single_day_schedule_info["worksheet"]
        end_row = single_day_schedule_info["ending_cell"].row
        schedule = []

        # zaczynamy od drugiej kolumny planu (następna po grupach)
        schedule_starting_row = single_day_schedule_info["starting_cell"].row
        schedule_starting_column = single_day_schedule_info["starting_cell"].column + 1
        for schedule_hour in range(AMOUNT_OF_TIME_BLOCK_PER_DAY):
            current_row = schedule_starting_row
            while current_row < end_row:
                cell = self.get_cell(worksheet, row=current_row, column=schedule_starting_column + schedule_hour)
                if cell in single_day_schedule_info["excluded_cells"]:
                    log.debug(f"Current cell: {cell.coordinate} excluded")
                    current_row += 1
                    continue
                log.debug(f"Current cell: {cell.coordinate}")

                if cell.value:
                    schedule_block = self.get_module_info(single_day_schedule_info, cell)
                    schedule.append(schedule_block)
                    cells_to_exclude = worksheet.iter_cols(
                        min_col=schedule_block["starting_cell"].column,
                        max_col=schedule_block["ending_cell"].column,
                        min_row=schedule_block["starting_cell"].row,
                        max_row=schedule_block["ending_cell"].row
                    )
                    for column in cells_to_exclude:
                        single_day_schedule_info["excluded_cells"].extend(column)

                    current_row = schedule_block["ending_cell"].row + 1
                else:
                    current_row += 3
        return schedule

    def get_module_info(self, single_day_schedule_info: Dict[str, Any], starting_cell: Cell) -> Dict[str, Any]:
        log.debug(f"Getting module info for cell: {starting_cell.coordinate}")
        worksheet: Worksheet = single_day_schedule_info["worksheet"]
        starting_cell, ending_cell = self.get_module_range(single_day_schedule_info, starting_cell)
        start, end = self.get_module_times(single_day_schedule_info, starting_cell, ending_cell)
        name = self.get_module_name(starting_cell)
        _module_info = {
            "name": name,
            "type": self.get_module_type_from_name(name),
            "start": start,
            "end": end,
            "lecturers": self.get_module_lecturers(worksheet, starting_cell, ending_cell),
            "rooms": self.get_module_rooms(worksheet, starting_cell, ending_cell),
            "groups": self.get_module_groups(single_day_schedule_info["groups"], starting_cell, ending_cell),
            "colour": self.get_module_colour(starting_cell),
            "starting_cell": starting_cell,
            "ending_cell": ending_cell,
        }
        log.debug(f"Module info: {_module_info}")
        return _module_info

    def get_module_range(self, single_day_schedule_info: Dict[str, Any], starting_cell: Cell) -> Tuple[Cell, Cell]:
        log.debug(f"Getting module range [starting cell: {starting_cell.coordinate}]")
        worksheet = single_day_schedule_info["worksheet"]
        bg_colour = self.get_cell_colour(starting_cell)

        log.debug(f"Getting module last column")
        end_cell_column = starting_cell.column + 0
        while True:
            if end_cell_column > single_day_schedule_info["ending_cell"].column:
                raise ValidationError("Błąd podczas próby pobrania wielkości modułu")

            possible_end_of_module = self.get_cell(
                worksheet, row=starting_cell.row, column=end_cell_column
            )
            possible_start_of_next_module = self.get_cell(
                worksheet, row=starting_cell.row, column=possible_end_of_module.column + 1
            )
            log.debug(f"Current cell: {possible_end_of_module.coordinate}")
            if possible_end_of_module.border.right.style or possible_start_of_next_module.border.left.style:
                break

            colour_to_check = self.get_cell_colour(possible_start_of_next_module)
            if bg_colour != "00000000" and colour_to_check != bg_colour:
                break

            end_cell_column += 1

        log.debug(f"Getting module last row")
        end_cell_row = starting_cell.row + 2
        while True:
            if end_cell_row > single_day_schedule_info["ending_cell"].row:
                raise ValidationError(f"Błąd podczas próby pobrania wielkości modułu [{starting_cell.coordinate}]")

            possible_end_of_module = self.get_cell(
                worksheet, row=end_cell_row, column=starting_cell.column
            )
            possible_start_of_next_module = self.get_cell(
                worksheet, row=possible_end_of_module.row + 1, column=starting_cell.column
            )
            log.debug(f"Current cell: {possible_end_of_module.coordinate}")
            if possible_end_of_module.border.bottom.style or possible_start_of_next_module.border.top.style:
                break

            bg_colour_to_check = self.get_cell_colour(possible_start_of_next_module)
            if bg_colour not in ["00000000", "FFFFFF00"] and bg_colour_to_check not in [bg_colour, "FFFFFF00"]:
                break

            end_cell_row += 3

        return starting_cell, self.get_cell(worksheet, row=end_cell_row, column=end_cell_column)

    @classmethod
    def get_module_name(cls, starting_cell: Cell) -> str:
        log.debug("Getting module name")
        if name := starting_cell.value:
            return str(name).strip()
        raise ValidationError(f"Nie można pobrać nazwy modułu [{starting_cell.coordinate}]")

    @classmethod
    def get_module_type_from_name(cls, module_name: Optional[str]) -> str:
        log.debug("Getting module type from name")
        if module_name:
            module_name = module_name.replace(" ", "")
            for type_tuple in MODULE_TYPES_TUPLE:
                if type_tuple[0] in module_name:
                    return type_tuple[1]

        return DEFAULT_MODULE_TYPE

    @classmethod
    def get_module_times(
        cls, single_day_schedule_info: Dict[str, Any], starting_cell: Cell, ending_cell: Cell
    ) -> Tuple[datetime, datetime]:
        log.debug("Getting module times")
        schedule_starting_cell: Cell = single_day_schedule_info["starting_cell"]
        start_hour = starting_cell.column - schedule_starting_cell.column
        end_hour = start_hour + (ending_cell.column - starting_cell.column)
        return (
            datetime.combine(single_day_schedule_info["date"], HOUR_BLOCK_START_TIMESPANS[start_hour][0]),
            datetime.combine(single_day_schedule_info["date"], HOUR_BLOCK_START_TIMESPANS[end_hour][1])
        )

    @classmethod
    def get_module_groups(cls, group_list: List[Dict[str, Any]], starting_cell: Cell, ending_cell: Cell) -> List[str]:
        log.debug("Getting module groups")
        filtered_groups = filter(
            lambda g: g["start_row"] <= ending_cell.row and g["end_row"] >= starting_cell.row, group_list
        )
        return [group["name"] for group in filtered_groups]

    @classmethod
    def get_module_lecturers(cls, worksheet: Worksheet, starting_cell: Cell, ending_cell: Cell) -> List[Dict[str, Any]]:
        log.debug("Getting module lecturers")
        lecturers = []
        lecturer_rows = worksheet.iter_rows(
            min_col=starting_cell.column,
            max_col=starting_cell.column,
            min_row=starting_cell.row + 1,
            max_row=ending_cell.row
        )
        for row in lecturer_rows:
            cell = row[0]
            lecturer_name: Optional[str] = f"{cell.value or ''}"
            lecturer_name = lecturer_name.replace("/", ",")
            if not lecturer_name or any(char.isdigit() for char in lecturer_name):
                return lecturers

            for lecturer in lecturer_name.split(","):
                # pobieramy salę przypisaną do prowadzącego (sala obok jego nazwiska)
                # lub jeśli nie ma jej podanej bierzemy domyślną salę z dolnego prawego rogu bloku
                lecturer_room = (
                    cls.get_cell(worksheet, row=cell.row, column=ending_cell.column).value
                    or ending_cell.value
                    or "brak"
                )
                lecturers.append(
                    {
                        "name": lecturer.strip(),
                        "room": f"{lecturer_room}".strip()
                    }
                )
        return lecturers

    @classmethod
    def get_module_rooms(cls, worksheet: Worksheet, starting_cell: Cell, ending_cell: Cell) -> List[str]:
        log.debug("Getting module rooms")
        rooms = []
        room_rows = worksheet.iter_rows(
            min_col=ending_cell.column,
            max_col=ending_cell.column,
            min_row=starting_cell.row + 1,
            max_row=ending_cell.row
        )
        for row in room_rows:
            cell = row[0]
            room_name: str = f"{cell.value or ''}"
            if not room_name:
                continue
            rooms.append(room_name)
        return rooms

    def get_cell_colour(self, cell: Cell) -> str:
        colour = "00000000"
        if cell.fill.fgColor.type == "theme":
            theme = cell.fill.fgColor.theme
            tint = cell.fill.fgColor.tint
            colour = theme_and_tint_to_rgb(self.current_workbook, theme, tint)
        elif cell.fill.fgColor.type == "rgb":
            colour = cell.fill.fgColor.rgb
        return colour

    def get_module_colour(self, starting_cell: Cell) -> str:
        log.debug("Getting module colour")
        try:
            colour = self.get_cell_colour(starting_cell)
            if len(colour) == 6:
                return f"#00{colour}"
            if len(colour) == 8:
                return f"#{colour}"
            log.warning(f"Wrong colour palette {colour} [{starting_cell.coordinate}]")
            return ""
        except Exception:
            log.error(f"Wrong colour palette [{starting_cell.coordinate}]")
            return ""

    @classmethod
    def find_lower_boundary_of_column(cls, worksheet: Worksheet, column: int) -> Cell:
        log.debug(f"Looking for lower boundary of column [{column}]")
        for row in range(worksheet.max_row, 1, -1):
            cell = cls.get_cell(worksheet=worksheet, row=row, column=column)
            if cell.border.top != Side() or cell.border.bottom != Side():
                return cell

    @classmethod
    def get_cell_with_value(
        cls, worksheet: Worksheet, value: Any, _range: Optional[Tuple[Tuple, Tuple]] = None, exact: bool = False
    ) -> Optional[Cell]:
        start_row, start_col, end_row, end_col = 1, 1, worksheet.max_row, worksheet.max_column
        if _range:
            start_row, start_col = _range[0]
            end_row, end_col = _range[1]

        for row in range(start_row, end_row):
            for column in range(start_col, end_col):
                cell: Cell = cls.get_cell(worksheet=worksheet, row=row, column=column)
                cell_value = cell.value
                if exact:
                    if cell_value == value:
                        return cell
                elif f"{value}" in f"{cell_value}":
                    return cell
        return None


@dataclass
class ScheduleService:
    excel_schedule_service: ExcelScheduleService

    @transaction.atomic
    def update_schedule_from_excel(self, schedule: Schedule) -> Dict[str, Any]:
        result = {
            "replaced_blocks": [],
            "added_blocks": [],
            "added_groups": [],
            "added_lecturers": [],
            "added_rooms": [],
            "added_courses": [],
            "errors": []
        }
        Schedule.objects.filter(id=schedule.id).update(progress=0, status="ONGOING")

        excel_schedule_info = self.excel_schedule_service.init_excel_worksheet(
            path=schedule.file.path,
            worksheet=schedule.worksheet_name,
            year=schedule.year,
            month=schedule.month
        )
        Schedule.objects.filter(id=schedule.id).update(progress=20)

        progress_step_par_day = 80 / len(excel_schedule_info["schedule_days"])
        for day_schedule_info in excel_schedule_info["schedule_days"]:
            _date = day_schedule_info['date']
            if error := day_schedule_info.get("error", None):
                log.error(f"{_date}: bład ({error})")
                continue

            group_names = [group["name"] for group in day_schedule_info["groups"]]
            active_groups_in_schedule = Group.objects.filter(name__in=group_names)
            replaced_schedule_blocks = ScheduleBlock.objects.filter(
                groups__in=active_groups_in_schedule,
                start__date=_date,
                end__date=_date
            )
            result["replaced_blocks"].extend(replaced_schedule_blocks)

            added_blocks = 0
            added_groups = 0
            added_lecturers = 0
            added_rooms = 0
            added_courses = 0
            for _block in self.excel_schedule_service.get_schedule_for_single_day(day_schedule_info):
                try:
                    schedule_block = ScheduleBlock(
                        is_public=False,
                        course_name=_block["name"],
                        course=None,
                        start=_block["start"],
                        end=_block["end"],
                        type=_block["type"],
                        colour=_block["colour"],
                    )
                    course, course_created = Course.objects.get_or_create(name=_block["name"])
                    if course_created:
                        django_log_action(user=schedule.creator, obj=course, action_flag=ADDITION)
                        added_courses += 1
                        result["added_courses"].append(course)

                    schedule_block.course = course

                    schedule_block_groups = []
                    for _group in _block["groups"]:
                        group, group_created = Group.objects.get_or_create(name=_group)
                        if group_created:
                            django_log_action(user=schedule.creator, obj=group, action_flag=ADDITION)
                            added_groups += 1
                            result["added_groups"].append(group)

                        schedule_block_groups.append(group)

                    schedule_block_rooms = []
                    for _room in _block["rooms"]:
                        room, room_created = Room.objects.get_or_create(name=_room)
                        if room_created:
                            django_log_action(user=schedule.creator, obj=room, action_flag=ADDITION)
                            added_rooms += 1
                            result["added_rooms"].append(room)

                        schedule_block_rooms.append(room)

                    schedule_block_lecturers = []
                    for _lecturer in _block["lecturers"]:
                        lecturer, lecturer_created = Lecturer.objects.get_or_create(last_name=_lecturer["name"])
                        if lecturer_created:
                            django_log_action(user=schedule.creator, obj=lecturer, action_flag=ADDITION)
                            added_lecturers += 1
                            result["added_lecturers"].append(lecturer)

                        room, room_created = Room.objects.get_or_create(name=_lecturer["room"])
                        # jakby sala nie była stworzona w poprzednim forze
                        if room_created:
                            django_log_action(user=schedule.creator, obj=room, action_flag=ADDITION)
                            added_rooms += 1
                            result["added_rooms"].append(room)

                        schedule_block_lecturers.append(
                            LecturerScheduleBlockThrough(schedule_block=schedule_block, lecturer=lecturer, room=room)
                        )

                    schedule_block.save()
                    schedule_block.rooms.set(schedule_block_rooms)
                    schedule_block.groups.set(schedule_block_groups)
                    LecturerScheduleBlockThrough.objects.bulk_create(schedule_block_lecturers)
                    added_blocks += 1
                    result["added_blocks"].append(schedule_block)
                    django_log_action(user=schedule.creator, obj=schedule_block, action_flag=ADDITION)
                except IntegrityError:
                    result["errors"].append(
                        f"{_block['starting_cell'].coordinate}:{_block['ending_cell'].coordinate} Błąd odczytu wagonika"
                    )
                except Exception as ex:
                    log.error(ex)
                    result["errors"].append(
                        f"{_block['starting_cell'].coordinate}:{_block['ending_cell'].coordinate} Nieznany błąd"
                    )

            log_dict = {
                "wagoniki": added_blocks,
                "grupy": added_groups,
                "sale": added_rooms,
                "prowadzący": added_lecturers,
                "kursy": added_courses
            }
            log.debug(f"{_date}: zamieniono {replaced_schedule_blocks.count()} / nowe {log_dict}")
            Schedule.objects.filter(id=schedule.id).update(progress=F("progress") + progress_step_par_day)

        Schedule.objects.filter(id=schedule.id).update(progress=100, status="FINISHED")

        schedule.refresh_from_db()
        schedule.replaced_schedule_blocks.set(result["replaced_blocks"])
        schedule.schedule_blocks.set(result["added_blocks"])
        schedule.rooms.set(result["added_rooms"])
        schedule.courses.set(result["added_courses"])
        schedule.groups.set(result["added_groups"])
        schedule.lecturers.set(result["added_lecturers"])
        schedule.errors = result["errors"]
        schedule.save()

        return result

    @transaction.atomic
    def publicate_schedule(self, schedule: Schedule) -> None:
        if schedule.status not in ["FINISHED", "REVERTED"]:
            raise ValidationError("Plan nie został pobrany z pliku")

        schedule.replaced_schedule_blocks.update(is_public=False)
        schedule.schedule_blocks.all().update(is_public=True)
        schedule.lecturers.all().update(is_public=True)
        schedule.rooms.all().update(is_public=True)
        schedule.courses.all().update(is_public=True)
        schedule.groups.all().update(is_public=True)
        schedule.status = "PUBLICATED"
        schedule.save()

    @transaction.atomic
    def revert_schedule_publication(self, schedule: Schedule) -> None:
        if schedule.status != "PUBLICATED":
            raise ValidationError("Plan nie został opublikowany")

        schedule.replaced_schedule_blocks.update(is_public=True)
        schedule.schedule_blocks.all().update(is_public=False)
        schedule.lecturers.all().update(is_public=False)
        schedule.rooms.all().update(is_public=False)
        schedule.courses.all().update(is_public=False)
        schedule.groups.all().update(is_public=False)
        schedule.status = "REVERTED"
        schedule.save()
